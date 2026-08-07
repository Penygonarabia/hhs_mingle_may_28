# -*- coding: utf-8 -*-
import os
import sys
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import odoo
from odoo.api import Environment
from odoo.modules.registry import Registry

def format_float_hours(val):
    if not val:
        return "0:00"
    h = int(val)
    m = int(round((val - h) * 60))
    if m >= 60:
        h += 1
        m -= 60
    return f"{h}:{m:02d}"

def generate_report():
    db_name = "dbprod"
    registry = Registry(db_name)
    with registry.cursor() as cr:
        env = Environment(cr, odoo.SUPERUSER_ID, {})
        from odoo.addons.pbi_service_dashboards.controllers.service_config import BOARDS
        from odoo.addons.pbi_service_dashboards.controllers import service_sql
        from odoo.addons.pbi_service_dashboards.controllers.service_sql import (
            ParamBinder, _source_ctx, compile_domain, _technician_guard_clause, _promoter_guard_clause
        )

        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Remove default sheet

        # Styling Definitions
        title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        section_font = Font(name="Calibri", size=11, bold=True, color="1B365D")
        sub_section_font = Font(name="Calibri", size=10, bold=True, color="2B579A")
        header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        raw_header_font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
        bold_font = Font(name="Calibri", size=10, bold=True)
        regular_font = Font(name="Calibri", size=10)
        raw_regular_font = Font(name="Calibri", size=9)

        title_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        section_fill = PatternFill(start_color="E8EEF5", end_color="E8EEF5", fill_type="solid")
        header_fill = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
        raw_header_fill = PatternFill(start_color="4A6984", end_color="4A6984", fill_type="solid")
        kpi_fill = PatternFill(start_color="F2F4F8", end_color="F2F4F8", fill_type="solid")
        zebra_fill = PatternFill(start_color="F9FAFC", end_color="F9FAFC", fill_type="solid")

        thin_side = Side(border_style="thin", color="D9D9D9")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_center = Alignment(horizontal="center", vertical="center")

        date_from = "2026-01-01 00:00:00"
        date_to = "2026-12-31 23:59:59"
        period_months = 12

        target_boards = [
            ("technician_analysis", "Technician Analysis"),
            ("service_analysis_technicians", "Service Analysis - Tech"),
            ("service_analysis_crd", "Service Analysis (CRD)"),
            ("service_analysis_crd_users", "Service Analysis - CRD Users"),
        ]

        INTERVAL_ENDPOINT_MAP = {
            "labor_hours": (
                "'Technician Reached'",
                "'Ready to Invoice / Closed'",
                "COALESCE(main.reached_ts, pt.technician_reached_date)",
                "COALESCE(main.ready_to_invoice_ts, pt.closed_datetime)"
            ),
            "travel_time_hours": (
                "'Technician Travel Started'",
                "'Technician Reached'",
                "COALESCE(main.travel_started_ts, pt.technician_started_date)",
                "COALESCE(main.reached_ts, pt.technician_reached_date)"
            ),
            "technician_travel_hours": (
                "'Technician Travel Started'",
                "'Technician Reached'",
                "COALESCE(main.travel_started_ts, pt.technician_started_date)",
                "COALESCE(main.reached_ts, pt.technician_reached_date)"
            ),
            "rtat_hours": (
                "'Service Created'",
                "'Closed'",
                "main.service_created_datetime",
                "COALESCE(main.closed_ts, pt.closed_datetime)"
            ),
            "scheduling_hours": (
                "'Service Created'",
                "'Scheduled'",
                "main.service_created_datetime",
                "COALESCE(main.scheduled_ts, pt.scheduled_date)"
            ),
            "job_closing_hours": (
                "'Ready to Invoice'",
                "'Closed'",
                "COALESCE(main.ready_to_invoice_ts, pt.closed_datetime)",
                "COALESCE(main.closed_ts, pt.job_card_completed_time)"
            ),
            "onhold_to_ready_hours": (
                "'On Hold - SP Req'",
                "'Parts Ready & Rescheduled'",
                "COALESCE(main.onhold_ts, pt.job_hold_date)",
                "main.parts_ready_ts"
            ),
            "ready_to_handover_hours": (
                "'Parts Ready & Rescheduled'",
                "'Spare Parts Handover'",
                "main.parts_ready_ts",
                "main.handover_ts"
            ),
            "quote_to_parts_added_hours": (
                "'Customer Need Quote'",
                "'Parts Added Req Service Charge'",
                "COALESCE(main.cst_need_quote_ts, pt.cstneedquote_date)",
                "main.parts_added_ts"
            ),
            "total_worked_hours": (
                "'Work Started'",
                "'Work Logged / Closed'",
                "main.service_created_datetime",
                "COALESCE(main.closed_ts, pt.closed_datetime)"
            ),
            "expected_completion_hours": (
                "'Assigned'",
                "'Expected Completion'",
                "main.service_created_datetime",
                "COALESCE(main.closed_ts, pt.closed_datetime)"
            ),
        }

        def get_raw_rows(board_cfg, item_cfg):
            cte_sql, field_map, cte_name = _source_ctx(item_cfg)
            main_field_map = {k: (f"main.{v}" if isinstance(v, str) else v) for k, v in field_map.items()}
            binder = ParamBinder()
            binder.params["date_from"] = date_from
            binder.params["date_to"] = date_to
            clauses = []
            clauses += compile_domain(board_cfg.scope, main_field_map, binder, odoo.SUPERUSER_ID, env)
            clauses += compile_domain(item_cfg.domain, main_field_map, binder, odoo.SUPERUSER_ID, env)
            guard = _technician_guard_clause(env, odoo.SUPERUSER_ID, binder, item_cfg.source)
            if guard: clauses.append(guard)
            p_guard = _promoter_guard_clause(env, odoo.SUPERUSER_ID, binder, item_cfg.source)
            if p_guard: clauses.append(p_guard)

            where = (" AND " + " AND ".join(clauses)) if clauses else ""
            m_field = item_cfg.measure.field if item_cfg.measure else "job_card_status"

            if m_field in INTERVAL_ENDPOINT_MAP:
                from_st_sql, to_st_sql, from_dt_sql, to_dt_sql = INTERVAL_ENDPOINT_MAP[m_field]
            else:
                from_st_sql = "'New / Created'"
                to_st_sql = "COALESCE(main.job_card_status, 'Closed')"
                from_dt_sql = "main.service_created_datetime"
                to_dt_sql = "COALESCE(main.closed_ts, pt.closed_datetime)"

            if item_cfg.source == "message_log":
                table_name = "mail_tracking_value / project_task"
                sql = f"""
                    WITH {cte_sql}
                    SELECT
                        main.task_id AS task_id,
                        COALESCE(pt.name, CONCAT('Task #', main.task_id::text)) AS task_name,
                        COALESCE(ru.name, 'Unassigned') AS user_name,
                        COALESCE(main.user_role, 'N/A') AS region_or_role,
                        'New' AS from_status,
                        COALESCE(main.ptml_final_taskstatus, 'Closed') AS to_status,
                        main.task_date AS from_datetime,
                        main.task_date AS to_datetime,
                        1.0 AS raw_value
                    FROM {cte_name} main
                    LEFT JOIN project_task pt ON pt.id = main.task_id
                    LEFT JOIN res_users u ON u.id = main.user_id
                    LEFT JOIN res_partner ru ON ru.id = u.partner_id
                    WHERE 1=1 {where}
                    ORDER BY main.task_date DESC
                    LIMIT 30
                """
            else:
                table_name = "project_task (jobcards CTE)"
                if m_field and m_field in field_map and m_field not in ("job_card_status", "action_status", "service_created_datetime"):
                    val_col = f"COALESCE(main.{m_field}::float, 0.0)"
                else:
                    val_col = "1.0"
                sql = f"""
                    WITH {cte_sql}
                    SELECT
                        main.task_id AS task_id,
                        COALESCE(pt.name, CONCAT('Task #', main.task_id::text)) AS task_name,
                        COALESCE(tech_p.name, 'Unassigned') AS user_name,
                        COALESCE(wcg.name, 'All Regions') AS region_or_role,
                        {from_st_sql} AS from_status,
                        {to_st_sql} AS to_status,
                        {from_dt_sql} AS from_datetime,
                        {to_dt_sql} AS to_datetime,
                        {val_col} AS raw_value
                    FROM {cte_name} main
                    LEFT JOIN project_task pt ON pt.id = main.task_id
                    LEFT JOIN res_users tech_u ON tech_u.id = main.technician_id
                    LEFT JOIN res_partner tech_p ON tech_p.id = tech_u.partner_id
                    LEFT JOIN work_center_group wcg ON wcg.id = main.work_center_group_id
                    WHERE 1=1 {where}
                    ORDER BY main.service_created_datetime DESC
                    LIMIT 30
                """

            env.cr.execute(sql, binder.params)
            rows = env.cr.dictfetchall()
            return table_name, m_field, rows

        for board_key, sheet_title in target_boards:
            board_cfg = BOARDS.get(board_key)
            if not board_cfg:
                continue

            ws = wb.create_sheet(title=sheet_title)
            ws.views.sheetView[0].showGridLines = True

            current_row = 1

            # Title Banner
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row+1, end_column=12)
            title_cell = ws.cell(row=current_row, column=1, value=f"SERVICE DASHBOARDS AUDIT & TALLY REPORT: {board_cfg.title}")
            title_cell.font = title_font
            title_cell.fill = title_fill
            title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

            for r in range(current_row, current_row+2):
                for c in range(1, 13):
                    ws.cell(row=r, column=c).fill = title_fill

            current_row += 3

            # Board Metadata Block
            ws.cell(row=current_row, column=1, value="Dashboard Scope:").font = bold_font
            ws.cell(row=current_row, column=2, value=board_cfg.title).font = regular_font
            ws.cell(row=current_row, column=4, value="Period:").font = bold_font
            ws.cell(row=current_row, column=5, value=f"{date_from[:10]} to {date_to[:10]}").font = regular_font
            current_row += 2

            # Section 1: KPI Summary Tiles
            ws.cell(row=current_row, column=1, value="1. KEY PERFORMANCE INDICATORS (KPIs)").font = section_font
            current_row += 1

            kpi_items = [item for item in board_cfg.items if item.type in ("kpi_single", "kpi_dual")]
            if kpi_items:
                headers = ["KPI Metric Name", "Value / Count", "Secondary Metric / Value 2", "Format", "Calculation Summary"]
                for col_num, h_text in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=h_text)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = thin_border
                current_row += 1

                for item in kpi_items:
                    kpi_data = service_sql.run_kpi(env, odoo.SUPERUSER_ID, board_cfg, item, date_from, date_to, period_months)
                    val = kpi_data.get("value", 0)
                    val2 = kpi_data.get("value2")

                    val_str = format_float_hours(val) if item.measure and item.measure.field.endswith("_hours") else val
                    val2_str = format_float_hours(val2) if (val2 is not None and item.measure_2 and item.measure_2.field.endswith("_hours")) else (val2 if val2 is not None else "-")

                    row_data = [
                        item.name,
                        val_str,
                        val2_str,
                        item.type,
                        item.info or f"{item.name} metric calculation"
                    ]

                    for col_num, cell_val in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_num, value=cell_val)
                        cell.font = regular_font
                        cell.border = thin_border
                        cell.fill = kpi_fill
                        if col_num in (2, 3):
                            cell.alignment = align_right
                            if isinstance(cell_val, (int, float)):
                                cell.number_format = "#,##0.00" if isinstance(cell_val, float) else "#,##0"
                        else:
                            cell.alignment = align_left
                    current_row += 1
                current_row += 2

            # Section 2: Chart Breakdown Summary & Raw Data
            ws.cell(row=current_row, column=1, value="2. CHART BREAKDOWNS & UNDERLYING RAW DATA").font = section_font
            current_row += 1

            chart_items = [item for item in board_cfg.items if item.type in ("bar", "pie")]
            for item in chart_items:
                ws.cell(row=current_row, column=1, value=f"Chart: {item.name}").font = sub_section_font
                current_row += 1

                # Chart Aggregated Summary Table
                breakdown = service_sql.run_breakdown(env, odoo.SUPERUSER_ID, board_cfg, item, date_from, date_to, [], period_months) or []

                if item.measure_2:
                    headers = ["Category / Group", item.series_labels[0] if item.series_labels else "Series 1", item.series_labels[1] if item.series_labels else "Series 2"]
                else:
                    headers = ["Category / Group", "Aggregated Chart Value"]

                for col_num, h_text in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=h_text)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = align_center
                    cell.border = thin_border
                current_row += 1

                if not breakdown:
                    cell = ws.cell(row=current_row, column=1, value="No summary data records available for selected period.")
                    cell.font = regular_font
                    cell.border = thin_border
                    current_row += 1
                else:
                    for idx, row in enumerate(breakdown):
                        label = row.get("label", "")
                        v1 = row.get("value", 0)
                        v2 = row.get("value2")

                        is_hours = (item.measure and item.measure.field.endswith("_hours"))
                        is_pct = (item.measure and item.measure.field.endswith("_pct"))

                        v1_display = format_float_hours(v1) if is_hours else (f"{v1:.2f}%" if is_pct else v1)

                        if item.measure_2:
                            is_hours_2 = (item.measure_2 and item.measure_2.field.endswith("_hours"))
                            v2_display = format_float_hours(v2) if is_hours_2 else v2
                            row_vals = [label, v1_display, v2_display]
                        else:
                            row_vals = [label, v1_display]

                        for col_num, cell_val in enumerate(row_vals, 1):
                            cell = ws.cell(row=current_row, column=col_num, value=cell_val)
                            cell.font = regular_font
                            cell.border = thin_border
                            if idx % 2 == 1:
                                cell.fill = zebra_fill
                            if col_num > 1:
                                cell.alignment = align_right
                                if isinstance(cell_val, (int, float)):
                                    cell.number_format = "#,##0.00" if isinstance(cell_val, float) else "#,##0"
                            else:
                                cell.alignment = align_left
                        current_row += 1
                current_row += 1

                # Chart Raw Data Table (With From Status, To Status, From Datetime, To Datetime & Raw Float/Val)
                tbl_name, target_field, raw_rows = get_raw_rows(board_cfg, item)

                ws.cell(row=current_row, column=1, value=f"  └─ Raw Audit Records for '{item.name}' (Table: {tbl_name} | Target Field: {target_field})").font = bold_font
                current_row += 1

                raw_headers = [
                    "Table Name", "Target Field Name", "Task ID", "Task Name / Number",
                    "User / Technician", "Region / Work Center",
                    "From Status", "To Status",
                    "From Datetime", "To Datetime",
                    "Raw Float/Val", "Formatted Val"
                ]
                for col_num, h_text in enumerate(raw_headers, 1):
                    cell = ws.cell(row=current_row, column=col_num, value=h_text)
                    cell.font = raw_header_font
                    cell.fill = raw_header_fill
                    cell.alignment = align_center
                    cell.border = thin_border
                current_row += 1

                if not raw_rows:
                    cell = ws.cell(row=current_row, column=1, value="No raw database records found for this domain.")
                    cell.font = raw_regular_font
                    cell.border = thin_border
                    current_row += 1
                else:
                    for r_idx, r_row in enumerate(raw_rows):
                        raw_val = r_row.get("raw_value", 0.0)
                        is_hrs = target_field.endswith("_hours")
                        is_p = target_field.endswith("_pct")
                        fmt_val = format_float_hours(raw_val) if is_hrs else (f"{raw_val:.2f}%" if is_p else raw_val)

                        f_dt = r_row.get("from_datetime")
                        f_dt_str = f_dt.strftime("%Y-%m-%d %H:%M") if isinstance(f_dt, (datetime.datetime, datetime.date)) else str(f_dt or "-")

                        t_dt = r_row.get("to_datetime")
                        t_dt_str = t_dt.strftime("%Y-%m-%d %H:%M") if isinstance(t_dt, (datetime.datetime, datetime.date)) else str(t_dt or "-")

                        raw_data_vals = [
                            tbl_name.split()[0],
                            target_field,
                            r_row.get("task_id"),
                            r_row.get("task_name"),
                            r_row.get("user_name"),
                            r_row.get("region_or_role"),
                            r_row.get("from_status"),
                            r_row.get("to_status"),
                            f_dt_str,
                            t_dt_str,
                            raw_val,
                            fmt_val
                        ]

                        for col_num, cell_val in enumerate(raw_data_vals, 1):
                            cell = ws.cell(row=current_row, column=col_num, value=cell_val)
                            cell.font = raw_regular_font
                            cell.border = thin_border
                            if r_idx % 2 == 1:
                                cell.fill = zebra_fill
                            if col_num in (3, 11, 12):
                                cell.alignment = align_right
                                if isinstance(cell_val, (int, float)):
                                    cell.number_format = "#,##0.00" if isinstance(cell_val, float) else "#,##0"
                            else:
                                cell.alignment = align_left
                        current_row += 1

                current_row += 2

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row in (1, 2):
                        continue
                    val_str = str(cell.value or '')
                    if '\n' in val_str:
                        val_str = max(val_str.split('\n'), key=len)
                    max_len = max(max_len, len(val_str))
                ws.column_dimensions[col_letter].width = max(max_len + 4, 14)

        # Tab 5: Formulas & Reference
        ws_ref = wb.create_sheet(title="Formulas & Reference")
        ws_ref.views.sheetView[0].showGridLines = True

        ws_ref.cell(row=1, column=1, value="SERVICE DASHBOARDS MASTER FORMULA & SCHEMA GLOSSARY").font = section_font
        ws_ref.cell(row=2, column=1, value="Detailed table names, target field names, status transition endpoints, SQL calculations, and business definitions for all 4 dashboards.").font = regular_font

        headers = ["Dashboard", "Chart / Metric Name", "Table Name", "Target Field Name", "From Status -> To Status", "Calculation & Business Formula"]
        for col_num, h_text in enumerate(headers, 1):
            cell = ws_ref.cell(row=4, column=col_num, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

        ref_data = [
            ("Technician Analysis", "Technician Closed Job Cards", "project_task", "job_card_status", "Open -> Closed", "COUNT(job cards WHERE state = 'Closed') grouped by Technician."),
            ("Technician Analysis", "Technician Jobs - Average RTAT", "project_task", "rtat_hours", "Service Created -> Closed", "AVG(rtat_hours) for Closed job cards (Closed Date - Created Date in float hours)."),
            ("Technician Analysis", "Technician Utilization", "project_task", "utilization_pct", "Reached -> Closed & Travel Started -> Reached", "((SUM(Labor Hours) + SUM(Travel Hours)) / (176 hrs * Months)) * 100 per Technician."),
            ("Technician Analysis", "Technician Labor Hours", "project_task", "labor_hours", "Technician Reached -> Ready to Invoice / Closed", "SUM(hours elapsed from Technician Reached at site to Job Completion)."),
            ("Technician Analysis", "Technician Travel Hours", "project_task", "travel_time_hours", "Travel Started -> Technician Reached", "SUM(hours elapsed from Travel Started to Technician Reached at site)."),
            ("Technician Analysis", "Employee Performance - Actual Hours", "project_task", "total_worked_hours", "Work Started -> Closed", "SUM(actual labor hours logged on job card work logs) grouped by Region."),
            ("Technician Analysis", "Employee Performance - Estimated vs Actual", "project_task", "expected_completion_hours, total_worked_hours", "Assigned -> Closed", "SUM(assigned estimated completion hours) vs SUM(actual logged work hours)."),
            ("Service Analysis - Tech", "My Closed Job Cards", "project_task", "job_card_status", "Open -> Closed", "COUNT(Closed job cards) assigned to current technician grouped by month."),
            ("Service Analysis - Tech", "My Labor Hours", "project_task", "labor_hours", "Technician Reached -> Ready to Invoice", "SUM(labor_hours) assigned to current technician grouped by month."),
            ("Service Analysis (CRD)", "Scheduling Performance", "project_task", "scheduling_hours", "Service Created -> Scheduled", "AVG(hours elapsed from Service Created Date to Scheduled Date) per Coordinator."),
            ("Service Analysis (CRD)", "Job Closing Performance", "project_task", "job_closing_hours", "Ready to Invoice -> Closed", "AVG(hours elapsed from Ready to Invoice to Closed Date) per Coordinator."),
            ("Service Analysis (CRD)", "Total Closed Job Cards", "project_task", "closed_by_uid", "Open -> Closed", "COUNT(Closed job cards) grouped by closing Coordinator."),
            ("Service Analysis - CRD Users", "Tasks - Month wise", "mail_tracking_value", "task_date", "New -> Final Status", "COUNT(tasks) tracked in status transition logs grouped by month."),
            ("Service Analysis - CRD Users", "My Scheduling Performance", "project_task", "scheduling_hours", "Service Created -> Scheduled", "AVG(scheduling_hours) for current coordinator's scheduled cards grouped by month."),
            ("Service Analysis - CRD Users", "My Job Closing Performance", "project_task", "job_closing_hours", "Ready to Invoice -> Closed", "AVG(job_closing_hours) for current coordinator's closed cards grouped by month."),
        ]

        for r_idx, r_data in enumerate(ref_data, 5):
            for c_idx, c_val in enumerate(r_data, 1):
                cell = ws_ref.cell(row=r_idx, column=c_idx, value=c_val)
                cell.font = regular_font
                cell.border = thin_border
                if r_idx % 2 == 1:
                    cell.fill = zebra_fill
                cell.alignment = align_left

        for col in ws_ref.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_ref.column_dimensions[col_letter].width = max(max_len + 4, 18)

        output_path = "/tmp/Service_Dashboards_Tally_Report.xlsx"
        wb.save(output_path)
        print(f"Report successfully saved to {output_path}")

if __name__ == "__main__":
    generate_report()
