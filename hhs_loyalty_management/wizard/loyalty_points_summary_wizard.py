# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
from datetime import date
from dateutil.relativedelta import relativedelta

import io
import base64
import re

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None


class LoyaltyPointsSummaryWizard(models.TransientModel):
    """
    Wizard to collect filter parameters and generate a
    Loyalty Points Summary Excel report.
    Uses a single SQL aggregation query for performance with
    large datasets (50 000+ customers / 500 000+ transactions).
    """
    _name = 'loyalty.points.summary.wizard'
    _description = 'Loyalty Points Summary Wizard'

    # ------------------------------------------------------------------
    # Fields
    # ------------------------------------------------------------------

    from_date = fields.Date(
        string='From Date',
        required=True,
        default=lambda self: date.today().replace(day=1),
    )

    to_date = fields.Date(
        string='To Date',
        required=True,
        default=lambda self: (
            date.today().replace(day=1) + relativedelta(months=1, days=-1)
        ),
    )

    customer_ids = fields.Many2many(
        'res.partner',
        'loyalty_summary_wiz_partner_rel',
        'wizard_id',
        'partner_id',
        string='Customers',
        domain=[('activate_loyalty_feature', '=', True)],
    )

    status_active = fields.Boolean(
        string='Active',
        default=True,
    )
    status_inactive = fields.Boolean(
        string='Inactive',
        default=False,
    )

    show_redeemed_only = fields.Boolean(
        string='Redeemed Points',
        # string='Show > 0 Redeemed Points',
        default=False,
    )

    # ------------------------------------------------------------------
    # Constraints
    # ------------------------------------------------------------------

    @api.constrains('from_date', 'to_date')
    def _check_dates(self):
        for rec in self:
            if rec.from_date and rec.to_date and rec.from_date > rec.to_date:
                raise UserError(
                    _("To Date must be greater than or equal to From Date.")
                )

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _get_report_lines(self):
        """
        Run the SQL aggregation and return the processed rows.
        Shared by action_generate_excel (and any future action).
        """
        self.ensure_one()

        if self.from_date > self.to_date:
            raise UserError(
                _("To Date must be greater than or equal to From Date.")
            )

        # Step 1: Resolve partner IDs
        if self.customer_ids:
            partner_ids = self.customer_ids.ids
        else:
            domain = []
            if self.status_active and not self.status_inactive:
                domain = [('activate_loyalty_feature', '=', True)]
            elif self.status_inactive and not self.status_active:
                domain = [('activate_loyalty_feature', '=', False)]
            # If both or neither are checked, no filter is applied (show all)
            partners = self.env['res.partner'].sudo().search(domain)
            partner_ids = partners.ids

        if not partner_ids:
            raise UserError(_("No customers found for the selected criteria."))

        # Step 2: SQL aggregation
        query = """
            SELECT
                p.id                                            AS partner_id,
                p.ref                                           AS customer_code,
                p.name                                          AS customer_name,
                p.mobile                                        AS mobile,
                p.city                                          AS city,
                COALESCE(p.salesman_name, '')                   AS salesman,
                COALESCE(p.tier_name, '')                       AS tier_name,
                wcg.name                                        AS region,

                -- OPENING BALANCE: net points before from_date
                COALESCE(SUM(
                    CASE WHEN h.clph_date < %(from_date)s THEN
                        CASE WHEN h.clph_doctype = '02' THEN -(COALESCE(h.clph_regpoints, 0) + COALESCE(h.clph_bonuspoints, 0))
                             WHEN h.clph_doctype IN ('98', '97') THEN -(COALESCE(h.clph_regpoints, 0) + COALESCE(h.clph_bonuspoints, 0))
                             WHEN h.clph_doctype = '99' AND h.clph_adjtype = '-' THEN -(COALESCE(h.clph_regpoints, 0) + COALESCE(h.clph_bonuspoints, 0))
                             ELSE COALESCE(h.clph_regpoints, 0) + COALESCE(h.clph_bonuspoints, 0)
                        END
                    ELSE 0 END
                ), 0)                                           AS opening_balance,

                -- REGULAR POINTS earned in range
                COALESCE(SUM(
                    CASE WHEN h.clph_date BETWEEN %(from_date)s AND %(to_date)s THEN
                        CASE WHEN h.clph_doctype = '01' THEN COALESCE(h.clph_regpoints, 0)
                             WHEN h.clph_doctype = '99' AND h.clph_adjtype = '+' THEN COALESCE(h.clph_regpoints, 0)
                             WHEN h.clph_doctype = '99' AND h.clph_adjtype = '-' THEN -COALESCE(h.clph_regpoints, 0)
                             WHEN h.clph_doctype = '02' THEN -COALESCE(h.clph_regpoints, 0)
                             ELSE 0
                        END
                    ELSE 0 END
                ), 0)                                           AS regular_points,

                -- BONUS POINTS earned in range
                COALESCE(SUM(
                    CASE WHEN h.clph_date BETWEEN %(from_date)s AND %(to_date)s THEN
                        CASE WHEN h.clph_doctype = '01' THEN COALESCE(h.clph_bonuspoints, 0)
                             WHEN h.clph_doctype = '99' AND h.clph_adjtype = '+' THEN COALESCE(h.clph_bonuspoints, 0)
                             WHEN h.clph_doctype = '99' AND h.clph_adjtype = '-' THEN -COALESCE(h.clph_bonuspoints, 0)
                             WHEN h.clph_doctype = '02' THEN -COALESCE(h.clph_bonuspoints, 0)
                             ELSE 0
                        END
                    ELSE 0 END
                ), 0)                                           AS bonus_points,

                -- REDEEMED in range
                COALESCE(SUM(
                    CASE WHEN h.clph_date BETWEEN %(from_date)s AND %(to_date)s
                              AND h.clph_doctype = '98'
                         THEN COALESCE(h.clph_regpoints, 0) ELSE 0 END
                ), 0)                                           AS redeemed_points,

                -- EXPIRED in range
                COALESCE(SUM(
                    CASE WHEN h.clph_date BETWEEN %(from_date)s AND %(to_date)s
                              AND h.clph_doctype = '97'
                         THEN COALESCE(h.clph_regpoints, 0) ELSE 0 END
                ), 0)                                           AS expired_points,

                -- TOTAL PURCHASE (set to 0 - purchase price not stored in loyalty history)
                0                                               AS total_purchase

            FROM res_partner p
            LEFT JOIN customer_loyalty_points_history h
                   ON h.clph_cstid = p.id
            LEFT JOIN res_city ccm
                   ON ccm.id = p.customer_city_id
            LEFT JOIN work_center_location wcl
                   ON wcl.id = ccm.def_work_center_id
            LEFT JOIN work_center_group wcg
                   ON wcg.id = wcl.work_center_group_id
            WHERE p.id = ANY(%(partner_ids)s)
            GROUP BY
                p.id, p.ref, p.name, p.mobile, p.city,
                p.salesman_name, p.tier_name, wcg.name
            ORDER BY p.name ASC
        """

        self.env.cr.execute(query, {
            'from_date': self.from_date,
            'to_date': self.to_date,
            'partner_ids': partner_ids,
        })
        rows = self.env.cr.dictfetchall()

        # Step 3: Compute available_points
        report_lines = []
        for row in rows:
            row['available_points'] = (
                row.get('opening_balance', 0)
                + row['regular_points']
                + row['bonus_points']
                - row['redeemed_points']
                - row['expired_points']
            )
            report_lines.append(row)

        # Step 4: Optional filter
        if self.show_redeemed_only:
            report_lines = [r for r in report_lines if r['redeemed_points'] > 0]

        if not report_lines:
            raise UserError(
                _("No loyalty records found for the selected criteria.")
            )

        return report_lines

    # ------------------------------------------------------------------
    # Excel generation
    # ------------------------------------------------------------------

    def action_generate_excel(self):
        """Generate and download a styled .xlsx Loyalty Points Summary."""
        if not openpyxl:
            raise UserError(
                _("openpyxl is not installed. Please install it: pip install openpyxl")
            )

        self.ensure_one()
        report_lines = self._get_report_lines()

        # ---- Styles --------------------------------------------------
        header_font   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        header_fill   = PatternFill('solid', fgColor='1F3864')   # dark navy
        sub_fill_even = PatternFill('solid', fgColor='DCE6F1')   # light blue
        sub_fill_odd  = PatternFill('solid', fgColor='FFFFFF')
        total_font    = Font(name='Calibri', bold=True, size=11)
        total_fill    = PatternFill('solid', fgColor='1F3864')
        total_color   = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
        center        = Alignment(horizontal='center', vertical='center', wrap_text=True)
        right         = Alignment(horizontal='right',  vertical='center')
        left          = Alignment(horizontal='left',   vertical='center')

        thin = Side(style='thin', color='B0B0B0')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        num_fmt = '#,##0'

        # ---- Workbook ------------------------------------------------
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Loyalty Points Summary'
        ws.sheet_view.showGridLines = False

        # ---- Title rows ----------------------------------------------
        company  = self.env.company.name
        user     = self.env.user.name
        from_str = self.from_date.strftime('%d-%m-%Y') if self.from_date else ''
        to_str = self.to_date.strftime('%d-%m-%Y') if self.to_date else ''

        ws.merge_cells('A1:O1')
        t = ws['A1']
        t.value     = f'{company} — Loyalty Points Summary'
        t.font      = Font(name='Calibri', bold=True, size=14, color='1F3864')
        t.alignment = center

        ws.merge_cells('A2:O2')
        s = ws['A2']
        s.value     = f'Period: {from_str}  to  {to_str}    |    Generated by: {user}'
        s.font      = Font(name='Calibri', size=10, color='595959')
        s.alignment = center

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 18

        # ---- Column headers ------------------------------------------
        columns = [
            ('Sl #',              6),
            ('Region',        16),
            ('City',          14),
            ('Salesman',      18),
            ('Customer Code', 16),
            ('Customer Name', 26),
            ('Mobile #',        14),
            ('Tier Name',          12),
            ('Opening\nBalance',   14),
            ('Regular\nPoints',    14),
            ('Bonus\nPoints',      14),
            ('Redeemed',          13),
            ('Expired',           12),
            ('Available\nPoints', 15),
            ('Total Purchase\nPrice',   14),
        ]

        header_row = 4
        ws.row_dimensions[header_row].height = 32

        for col_idx, (label, width) in enumerate(columns, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=label)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = border
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # ---- Data rows -----------------------------------------------
        numeric_cols = {9, 10, 11, 12, 13, 14, 15}  # 1-based col indices

        # Helper to strip illegal XML/Excel control characters from strings
        def _clean(val):
            if not isinstance(val, str):
                return val
            # Remove characters that openpyxl/Excel cannot handle
            return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', val)

        for row_idx, line in enumerate(report_lines, start=1):
            excel_row = header_row + row_idx
            fill = sub_fill_even if row_idx % 2 == 0 else sub_fill_odd
            ws.row_dimensions[excel_row].height = 16

            values = [
                row_idx,
                _clean(line.get('region') or ''),
                _clean(line.get('city') or ''),
                _clean(line.get('salesman') or ''),
                _clean(line.get('customer_code') or ''),
                _clean(line.get('customer_name') or ''),
                _clean(line.get('mobile') or ''),
                _clean(line.get('tier_name') or ''),
                line.get('opening_balance', 0),
                line.get('regular_points', 0),
                line.get('bonus_points', 0),
                line.get('redeemed_points', 0),
                line.get('expired_points', 0),
                line.get('available_points', 0),
                line.get('total_purchase', 0),
            ]

            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=val)
                cell.fill   = fill
                cell.border = border
                if col_idx in numeric_cols:
                    cell.alignment    = right
                    cell.number_format = num_fmt
                elif col_idx == 1:
                    cell.alignment = center
                else:
                    cell.alignment = left

        # ---- Totals row ----------------------------------------------
        total_row = header_row + len(report_lines) + 1
        ws.row_dimensions[total_row].height = 18

        total_keys = [
            'opening_balance', 'regular_points', 'bonus_points',
            'redeemed_points', 'expired_points', 'available_points',
            'total_purchase',
        ]
        totals = {k: sum(r.get(k, 0) for r in report_lines) for k in total_keys}

        total_values = [
            'TOTALS', '', '', '', '', '', '', '',
            totals['opening_balance'],
            totals['regular_points'],
            totals['bonus_points'],
            totals['redeemed_points'],
            totals['expired_points'],
            totals['available_points'],
            totals['total_purchase'],
        ]

        for col_idx, val in enumerate(total_values, start=1):
            cell = ws.cell(row=total_row, column=col_idx, value=val)
            cell.font   = total_color
            cell.fill   = total_fill
            cell.border = border
            if col_idx in numeric_cols:
                cell.alignment    = right
                cell.number_format = num_fmt
            else:
                cell.alignment = center

        # Freeze panes below header
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

        # ---- Save to attachment and return download URL --------------
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        xlsx_data = base64.b64encode(buffer.read())

        filename = f'Loyalty_Points_Summary_{from_str}_to_{to_str}.xlsx'

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': xlsx_data,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'res_model': self._name,
            'res_id': self.id,
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }
